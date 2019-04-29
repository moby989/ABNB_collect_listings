#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Feb  2 20:18:55 2019

@author: moby
"""

from __future__ import print_function
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from datetime import date
import csv


class FileWriter(object):

    def __init__(self, data):

        self.data = data #amke sure the data is in proper format eg list of dict
#        self.fields=list(fields.keys())             
        self.today = str(date.today())
        self.scopes = ['https://www.googleapis.com/auth/drive']

    def output_file(self,out_format,file_name,folder_name = None):

        format = out_format
                
#        if format == 'json' or format is None:
#            import json
#            file_name = 'hotels-in-{country}.txt'.format(
#                country=self.country.replace(" ", "-"))
#            with open(file_name, 'w', encoding='utf-8') as outfile:
#                json.dump(list(self.data), outfile, indent=2, ensure_ascii=False)

        if format == 'excel':

            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            fieldnames = list(self.data[0].keys())
            
            for i in range (1,len(self.data[1:])+2): #iterating via rows
                for j in range (1,len(fieldnames)+1):  #iteration via coloumns
                    if i == 1:
                        ws.cell(row=i, column=j).value = fieldnames[j-1]
                    else:
                        ws.cell(row=i, column=j).value = self.data[1:][i-2].get(fieldnames[j-1])       
                                                                            
            file_to_create_name = '{name}{today}.xlsx'.format(name=file_name,today=self.today)
            
#            save_path = os.path.join(os.getcwd(),folder_name)

            complete_name = file_to_create_name
#            os.path.join(save_path, file_to_create_name)

            wb.save(complete_name)
#            self.file_uploadGDrive_token(complete_name)                            
                    
        else:

            file_to_create_name = '{name}{today}.csv'.format(name = file_name,today = self.today)

#            save_path = os.path.join(os.getcwd(),folder_name)
#            complete_name = os.path.join(save_path, file_to_create_name)

            with open(file_to_create_name, mode='w') as csv_file:

                fieldnames = self.data[0].keys()
                writer = csv.DictWriter(csv_file, fieldnames = fieldnames)
                writer.writeheader()
                writer.writerows(self.data[1:])
        
#        self.file_uploadGDrive_token(file_to_create_name)                            
        
        print ('File -> '+str(file_to_create_name)+' is saved')
        
        return file_to_create_name
            
                
    def upload_file_to_GoogleDrive_Oath(self, file_name, folder_name):

        """
        DEPRECIATED

        """
        
        id = '1mnTwirEO0jUXGoZgiD9AkUc5LfGaxrWZ' #ID of folder 'Scraped_data'     
    
#        drive = self.access_gDrive()

        g_login = GoogleAuth()  
        g_login.LocalWebserverAuth()
        drive = GoogleDrive(g_login)         

        #get list of folders in the main_folder to upload (id above, folder name 'Scraped_data')

        file_list = drive.ListFile({'q': "'1UN_U6lgrFFHQ1booBtB90FSkWM_ZnohO' in parents and trashed=false"}).GetList()
        
        folder_names = []

        for file in file_list:
        
            folder_names.append(file['title']) #collect names of folders into the list
        
        if folder_name not in folder_names: #check if the upload folder doesn't exist, if not exist create new
        
        #create folder to upload the files on GoogleDrive    

            work_folder = drive.CreateFile({'title': folder_name, \
                "parents":  [{"id": id}], \
                "mimeType": "application/vnd.google-apps.folder"})
            work_folder.Upload()
        
        file_list = drive.ListFile({'q': "'1UN_U6lgrFFHQ1booBtB90FSkWM_ZnohO' in parents and trashed=false"}).GetList()

        #get id of the folder to upload
        
        for file in file_list:
            
            if file['title'] == folder_name:
                
                folder_id = file['id']

        #upload file to the destination folder
            
        with open(file_name,"r") as file:
            file_drive = drive.CreateFile({'title':file_name,\
                "parents": [{"kind": "drive#fileLink","id": folder_id}]})
            file_drive.SetContentFile(file_name) 
            file_drive.Upload()
            print ("The file: " + file_name + " has been uploaded to GoogleDrive into the folder " +str(folder_name))    
            
    